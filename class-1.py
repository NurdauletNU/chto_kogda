from PyQt6.QtCore import QRect
from PyQt6.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QTextEdit, QGridLayout
import sys
import sqlite3
import openpyxl



conn=sqlite3.connect("database.db")
cursor=conn.cursor()

create_table_query= """ 
CREATE TABLE IF NOT EXIST users (
id INTEGER PRIMARY KEY,
    name TEXT,
    email TEXT
);
"""

cursor.execute(create_table_query)
conn.commit()

def import_data(file_path):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook.active

    insert_query="INSERT INTO users (name,email) VALUES (?,?)"
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name,email=row
        cursor.execute(insert_query,(name,email))
    conn.commit()


def export_data(file_path):
    select_query="SELECT name,email FROM users"
    cursor.execute(select_query)
    rows=cursor.fetchall()

    workbook=openpyxl.Workbook()
    sheet=workbook.active
    sheet.append("NAME", "Email")
    for row in rows:
        sheet.append(row)
    workbook.save(file_path)

import_file_path="data.xlsx"
export_file_path="export.xlsx"


import_data(import_file_path)
export_data(export_file_path)

conn.close()


