import json
import requests
import openpyxl



response=requests.get("https://jsonplaceholder.typicode.com/posts")
data=response.json()
print(type(data),len(data))



new_workbook=openpyxl.Workbook()
new_worksheet=new_workbook.active

# запись заголовков в эксель файл
titles=["userId", "id", "title", "body"]
index=0
for title in titles:
    index+=1
    new_worksheet.cell(row=1,column=index,value=title)

#new_workbook.save("new.xlsx")



# запись содержание под заголовок
index=1
for d in data:
    index+=1
    new_worksheet.cell(row=index,column=1,value=d["userId"])
    new_worksheet.cell(row=index,column=2,value=d["id"])
    new_worksheet.cell(row=index,column=3,value=d["title"])
    new_worksheet.cell(row=index,column=4,value=d["body"])

#new_workbook.save("new.xlsx")

