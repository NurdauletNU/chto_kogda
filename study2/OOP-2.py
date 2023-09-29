import datetime
import  json
import openpyxl

# 1. Реализовать класс для заданной структуры
# 2. Прочитать весь файл в оперативную память, в эту структуру
# 3. Реализовать возврат каждой сущности в json
# 4. Отфильтровать только людей в кадровом резерве
# 5. Добавить им всем БОНУС +15% к зарплате и выгрузить их
# 6. Отсортировать в порядке возрастания.

# Имя   Должность  Зарплата   Кадровый резерв   Дата устройство
# Вася  Инженер    90000       да                1 января 2023г



# 1. Реализовать класс для заданной структуры

class Person:
    additional_bonus=1.15


    def __repr__(self):
        return f"<Person> {self.name} {self.salary}"

    def __init__(self,name:str,posotion:str,salary:any,personal_register:str,date_device):
        self.name:str=name
        self.position:str=posotion
        self.salary:float=float(salary)
        self.personal_register:bool=str(personal_register).strip().lower()=="да"
        self.date_device: datetime.datetime=date_device

    # 3. Реализовать возврат каждой сущности в json

    def serialize_to_json(self):
        return {"name":self.name, "position":self.position,"salary":self.salary,"personal_register":self.personal_register
                "date_device":self.date_device}






# 2. Прочитать весь файл в оперативную память, в эту структуру

persons=[]
workbook=openpyxl.load_workbook("new.xlsx")
worksheet=workbook.active
for i in range(2,worksheet.max_row+1):
    person=Person(
        name=worksheet.cell(row=i,column=1).value,
        posotion=worksheet.cell(row=i,column=2).value,
        salary=worksheet.cell(row=i,column=3).value,
        personal_register=worksheet.cell(row=i,column=4).value,
        date_device=worksheet.cell(row=i,column=4).value,
    )

    persons.append(person)
#print(persons)
























