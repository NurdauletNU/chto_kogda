# Есть 3 excel-файла с данными в архиве. нужно:
# Записать все 3 файла в один новый общий на одной страницу
# Прочитать 3 файла -> записываем в массивы. Записываем массивы в один файл



from openpyxl import Workbook
from openpyxl import load_workbook


combined_workbook = Workbook()
combined_sheet = combined_workbook.active

files = ["Лист1.xlsx", "Лист2.xlsx", "Лист3.xlsx"]
columns = ['A', 'B', 'C']

for file, column in zip(files, columns):
    workbook = load_workbook(filename=file)
    sheet = workbook.active
    column_values = sheet[column]

    for value in column_values:
        combined_sheet[column + str(value.row)] = value.value


combined_workbook.save(filename="Общий.xlsx")
