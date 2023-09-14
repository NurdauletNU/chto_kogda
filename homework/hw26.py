# Доработать настольное приложение для ведения «списка кандидатов» на вакансии,
# добавить функционал с отправкой писем, дополнительными окнами и т.д

import sys
import sqlite3
import datetime
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt6.QtWidgets import QApplication, QWidget, QMessageBox, QGridLayout, QPushButton
from PyQt6 import uic
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLineEdit

class Ui(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = uic.loadUi("hw25.ui", self)
        self.ui.pushButton_save.clicked.connect(self.save_to_database)
        self.ui.pushButton_export.clicked.connect(self.export_from_database)
        self.show()



    def send_email(to_email, subject, message):
        smtp_server = 'smtp.example.com'
        smtp_port = 587
        smtp_username = 'your_username'
        smtp_password = 'your_password'
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_username, smtp_password)
        email_message = MIMEMultipart()
        email_message['From'] = smtp_username
        email_message['To'] = to_email
        email_message['Subject'] = subject
        email_message.attach(MIMEText(message, 'plain'))
        server.sendmail(smtp_username, to_email, email_message.as_string())
        server.quit()

    def save_to_database(self):
        try:
            # номер заявки
            number = self.lineEdit_number.text().strip()
            is_number = self.validate_number(value=number)
            if is_number is False:
                # self.app2 = Modal()  # всплывающее модальное окно
                # self.app2.show()
                QMessageBox.critical(self, "Ошибка", "Вы ввели неправильный номер заявки")
                return

            # сумма
            price = self.doubleSpinBox_price.value()
            # is_price = self.validate_price(value=price)
            # if is_number is False:
            #     QMessageBox.critical(self, "Ошибка", "Вы ввели неправильный сумму")
            #     return

            # TODO Запись в базу данных
            # print(number, price)

            # отправка в базу данных
            connection = sqlite3.connect("database_items.db")
            cursor = connection.cursor()

            # TODO SQL INJECTION ########################################
            # вредоносный SQL-код
            # number = '124314;drop table postgres;'
            query = f"""
INSERT INTO items (number, price)
VALUES ('{number}', '{price}')
"""
            # TODO SQL INJECTION ########################################

            # safe to SQL INJECTION
            # сама библиотека проверяет все переменные и удаляет возможный вредоносный код
            query = f"""
            INSERT INTO items (number, price)
            VALUES (?, ?)
            """
            cursor.execute(query, (number, price))
            connection.commit()
            #

            QMessageBox.information(self, "Успешно", "Заявка успешно добавлена в базу данных")
        except Exception as error:
            print(error)
            candidate_name = self.lineEdit_candidate_name.text()
            candidate_email = self.lineEdit_candidate_email.text()
            email_subject = 'Новый кандидат'
            email_message = f'Добрый день!\n\nНовый кандидат: {candidate_name}\nEmail: {candidate_email}'
            #send_email('recipient@example.com', email_subject, email_message)

            QMessageBox.information(self, "Успешно", "Заявка успешно добавлена в базу данных")
        except Exception as error:
            print(error)

    def export_from_database(self):
        try:
            
            rows: list[tuple[any]] = self.get_all_items()


            # создание в оперативной памяти excel файла и его открытие
            workbook: Workbook = openpyxl.Workbook()
            worksheet: Worksheet = workbook.active

            # запись заголовков в excel-файл
            for column_i, column in enumerate(["ИИН", "Номер заявки", "Сумма заявки"], 1):
                worksheet.cell(row=1, column=column_i, value=column)

            # запись всех данных в excel-файл
            for row_i, row in enumerate(rows, 2):
                for column_i, column in enumerate(row, 1):
                    worksheet.cell(row=row_i, column=column_i, value=column)

            workbook.save(f"экспорт_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx")
            QMessageBox.information(self, "Успешно", "Данные успешно экспортированы")
        except Exception as error:
            print(error)

    def validate_number(self, value: str) -> bool:
        """Проверка номера заявки число"""
        try:
            value = int(value)
            if value < 1:
                return False
        except Exception as error:
            print("Error: ", error)
            return False
        return True

    def validate_price(self):
        pass

    def get_all_items(self):
        connection = sqlite3.connect("database_items.db")
        cursor = connection.cursor()
        query1 = """
SELECT * FROM items;
"""
        query2 = """
INSERT INTO items (number, price)
VALUES ('134000', '666.67')
"""
        query3 = """
SELECT COUNT(*) FROM items;
"""
        cursor.execute(query1)
        # connection.commit()
        rows = cursor.fetchall()  # fetchone()
        print(rows)
        return rows

    class EditCandidateDialog(QDialog):
        def __init__(self):
            super().__init__()

            self.setWindowTitle('Редактировать кандидата')

            layout = QVBoxLayout()

            self.name_edit = QLineEdit()
            self.email_edit = QLineEdit()

            save_button = QPushButton('Сохранить')
            save_button.clicked.connect(self.save_candidate)

            layout.addWidget(self.name_edit)
            layout.addWidget(self.email_edit)
            layout.addWidget(save_button)

            self.setLayout(layout)

        def save_candidate(self):
            # Получите данные из полей редактирования и выполните необходимые операции сохранения
            candidate_name = self.name_edit.text()
            candidate_email = self.email_edit.text()

            # Закройте окно после сохранения
            self.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ui = Ui()
    sys.exit(app.exec())

    def database_postgre():
        """
        CREATE TABLE items (
        id SERIAL PRIMARY KEY,
        number BIGINT NOT NULL,
        price DECIMAL(10, 2) NOT NULL default = '0.0'
        );
        """

    def database_sqlite():
        # создание соединения с базой данных
        connection = sqlite3.connect("database_items.db")  # SELECT name FROM sqlite_master WHERE type='table';

        # создание "курсора" к базе данных
        cursor = connection.cursor()

        # Написание запроса для создания базы данных на SQL
        query = """
CREATE TABLE IF NOT EXISTS items 
(
id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
number INTEGER NOT NULL,
price REAL default '0.0'
)
"""
        # Выполнение запроса
        cursor.execute(query)

    class Modal(QWidget):
        def __init__(self, parent=None):
            super().__init__(parent)
            # self.setWindowModality(Qt.WindowModal)
            # self.setModal(True)
            self.resize(400, 400)
            self.setWindowTitle("Вы ввели неверный номер заявки")

            self.grid = QGridLayout(self)

            self.button = QPushButton()
            self.button.setText("ОК")

            self.grid.addWidget(self.button)