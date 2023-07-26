import sys
import sqlite3
from PyQt6.QtWidgets import QApplication, QMainWindow, QLabel, QPushButton, QFileDialog
from openpyxl import Workbook, load_workbook

def import_data(file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row[1:4])  # Наименование, Количество, Цена
    return data

def export_data(file_path):
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    cursor.execute("SELECT id, наименование, количество, цена FROM products")
    data = cursor.fetchall()
    conn.close()

    wb = Workbook()
    sheet = wb.active
    sheet.append(["id", "наименование", "количество", "цена"])
    for row in data:
        sheet.append(row)

    wb.save(file_path)

def import_button_clicked():
    file_dialog = QFileDialog()
    file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")
    file_dialog.exec()

    file_path = file_dialog.selectedFiles()[0]
    data = import_data(file_path)


def export_button_clicked():
    file_dialog = QFileDialog()
    file_dialog.setDefaultSuffix(".xlsx")
    file_dialog.setNameFilter("Excel Files (*.xlsx)")
    file_dialog.exec()

    file_path = file_dialog.selectedFiles()[0]
    export_data(file_path)


if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = QMainWindow()
    window.setWindowTitle("Экспорт-импорт данных")
    window.setGeometry(200, 200, 640, 480)

    label = QLabel("Выберите файл:", window)
    label.setGeometry(10, 10, 280, 30)

    import_button = QPushButton("Импортировать", window)
    import_button.setGeometry(10, 50, 130, 30)
    import_button.clicked.connect(import_button_clicked)

    export_button = QPushButton("Экспортировать", window)
    export_button.setGeometry(160, 50, 130, 30)
    export_button.clicked.connect(export_button_clicked)

    window.show()
    sys.exit(app.exec())
