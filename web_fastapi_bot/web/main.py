import aiohttp
from fastapi import FastAPI, Request, Depends
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
import openpyxl
import io
import aiosqlite
from starlette.responses import RedirectResponse

app = FastAPI()
templates = Jinja2Templates(directory="templates")  # Указываем директорию с шаблонами

@app.get("/")
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/submit")
async def submit(request: Request):
    form_data = await request.form()
    file = form_data["addition"]
    f = await file.read()
    workbook = openpyxl.load_workbook(filename=io.BytesIO(f), data_only=True)
    worksheet = workbook.active
    data = tuple(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column, values_only = True))
    print(data)
    clear_data = []
    for i in data:
        is_have_none = False
        for j in i:
            if j is None:
                is_have_none = True
        if not is_have_none:
            clear_data.append(i)
        # if all(j is not None for j in i):  # синтаксический сахар
        #     clear_data.append(i)
    print(clear_data)



    async with aiosqlite.connect("database.db") as db:
        await db.execute(''' 
CREATE TABLE IF NOT EXISTS items(
    id INTEGER PRIMARY KEY UNIQUE NOT NULL,
    name TEXT NOT NULL,
    description TEXT NOT NULL,
    price REAL default 0.0,
    quantity INTEGER default 0)''')
        await db.commit()

        it=clear_data[0]
        beaty = f"[{it[0]}] {it[1]} - {it[2][:30]} ... {it[3]}$ | {it[4]}"
        await db.execute('''
        INSERT INTO items(id, name, description, price, quantity)
        VALUES(:id, :name, :description, :price, :quantity)''',
        {"id": it[0], "name": it[1], "description": it[2], "price": it[3], "quantity": it[4]})
        await db.commit()

        bot_token = "6892706561:AAFTaxybOdKfN53-XlHDqqorqe_mYVnOn2Q"
        users = '465139386'
        async with aiohttp.ClientSession() as session:
            for user in [str(x).strip() for x in users.split(',')]:
                async with session.get(f'https://api.telegram.org/bot{bot_token}/sendMessage',
                                       params={'chat_id': user, 'text': beaty}) as response:
                    await response.json()


        return RedirectResponse(url="/")



